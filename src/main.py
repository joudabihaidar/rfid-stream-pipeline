import argparse
import logging
import sys
import openpyxl
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from .pipeline import run_pipeline_from_rows, flush_stale_buffers
from .tracking import EPCTracker, MovementTracker
from .aggregation import aggregate_window
from .ingestion import stream_rfid_excel
from .preprocessing import preprocess_row
from .database import (
    create_schema,
    ingest_raw_reads,
    ingest_single_row,
    write_event_to_db,
    clear_session_data,
    get_raw_rows,
    get_all_sessions,
    get_events_for_session,
    insert_session,
    insert_events,
    insert_anomalies,
    update_session_quality,
)
from analytics.anomaly import detect_anomalies
from analytics.quality import compute_data_quality

EXCEL_PATH   = PROJECT_ROOT / "data" / "raw_data.xlsx"
LOG_PATH     = PROJECT_ROOT / "data" / "pipeline.log"


# ═══════════════════════════════════════════════════════════════
# LOGGING SETUP
# ═══════════════════════════════════════════════════════════════

def setup_logging():
    """
    Configures logging to write to both the terminal and a log file.
    Terminal: INFO and above.
    Log file: DEBUG and above, timestamped.
    """
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler(LOG_PATH, encoding="utf-8"),
        ],
    )
    logging.getLogger("openpyxl").setLevel(logging.WARNING)


log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# DISPLAY
# ═══════════════════════════════════════════════════════════════

def print_events(events: List[Dict]):
    for e in events:
        tag   = f"[{e['event']:<16}]"
        t0    = str(e["t0"]) if e["t0"] is not None else "N/A"
        door  = e.get("door", "")
        zone  = (
            f"{e.get('from_zone','').split('__')[-1]}"
            f" → "
            f"{e.get('to_zone','').split('__')[-1]}"
            if e["event"] == "ZONE_TRANSITION" else ""
        )
        dwell = f"  dwell={e['dwell_time']}s" if e.get("dwell_time") is not None else ""
        note  = f"  {e['note']}"              if e.get("note")        else ""
        print(f"  {tag}  t0={t0:<4}  {door or zone}{dwell}{note}")


# ═══════════════════════════════════════════════════════════════
# BATCH MODE — process all sessions at full speed
# ═══════════════════════════════════════════════════════════════

def process_sheet(file_path: str, sheet_name: str) -> tuple:
    """
    Runs the full two-stage pipeline for a single session.
    Returns (events, anomalies). Returns ([], []) on failure.
    """
    log.info(f"[{sheet_name}] [1/4] Extracting raw RFID reads into database...")
    try:
        ingest_raw_reads(file_path, sheet_name)
    except Exception as e:
        log.error(f"[{sheet_name}] Ingestion failed: {e}")
        return [], []

    try:
        db_rows = get_raw_rows(sheet_name)
    except Exception as e:
        log.error(f"[{sheet_name}] DB read failed: {e}")
        return [], []

    if not db_rows:
        log.warning(f"[{sheet_name}] No rows found — skipping.")
        return [], []

    log.debug(f"[{sheet_name}] {len(db_rows)} raw reads loaded.")

    log.info(f"[{sheet_name}] [2/4] Running detection pipeline...")
    try:
        events = run_pipeline_from_rows(db_rows)
    except Exception as e:
        log.error(f"[{sheet_name}] Pipeline failed: {e}")
        return [], []

    if not events:
        log.warning(f"[{sheet_name}] No events detected.")
        return [], []

    log.info(f"[{sheet_name}] [3/4] Detecting anomalies...")
    try:
        anomalies = detect_anomalies(events, db_rows)
        quality   = compute_data_quality(events, db_rows)
    except Exception as e:
        log.error(f"[{sheet_name}] Anomaly detection failed: {e}")
        anomalies = []
        quality   = {}

    log.info(f"[{sheet_name}] [4/4] Storing results...")
    try:
        insert_session(sheet_name, events, has_anomaly=len(anomalies) > 0)
        insert_events(sheet_name, events)
        insert_anomalies(sheet_name, anomalies)
        update_session_quality(sheet_name, quality)
    except Exception as e:
        log.error(f"[{sheet_name}] DB write failed: {e}")
        return [], []

    log.info(f"[{sheet_name}] ✓ Done — {len(events)} events, {len(anomalies)} anomalies.")
    return events, anomalies


def run_all_sheets(file_path: str) -> Dict[str, List[Dict]]:
    """
    Batch mode: processes all sessions at full speed.
    """
    log.info(f"Opening workbook: {file_path}")
    try:
        wb          = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except FileNotFoundError:
        log.error(f"Excel file not found: {file_path}")
        return {}
    except Exception as e:
        log.error(f"Failed to open workbook: {e}")
        return {}

    log.info(f"Found {len(sheet_names)} sessions: {sheet_names}")
    all_results = {}

    for sheet_name in sheet_names:
        print(f"\n{'═'*60}")
        print(f"  Processing session: {sheet_name}")
        print(f"{'═'*60}")

        events, anomalies = process_sheet(str(file_path), sheet_name)

        if events:
            all_results[sheet_name] = events
            print()
            print_events(events)
        else:
            print(f"  ✗ Session skipped — see log for details.")

    return all_results


# ═══════════════════════════════════════════════════════════════
# STREAMING MODE — row by row, real-time simulation
# ═══════════════════════════════════════════════════════════════

def run_streaming(file_path: str, delay: float = 0.02):
    """
    Streaming mode: processes all sessions row by row with a delay.

    Uses stream_rfid_excel() — the same function used everywhere
    in the pipeline — as the data source. No logic is duplicated.

    For each row:
      1. ingest_single_row()   — writes raw read to DB immediately
      2. preprocess_row()      — clean and cast
      3. flush_stale_buffers() — check for confirmed events
      4. write_event_to_db()   — writes each event to DB immediately

    At the end of each session:
      5. detect_anomalies()    — runs on accumulated session events
      6. compute_data_quality() — ghost ratio + RSSI strength
      7. stores results to DB

    Trackers reset between sessions — each sheet is an independent
    test case, not a continuous recording. Shared trackers would
    cause state from session N to bleed into session N+1.

    WAL mode on SQLite allows the dashboard to read while this
    function is writing — no locking errors.

    Args:
        file_path: path to the Excel workbook
        delay:     seconds between rows (0.02 = 50 rows/second)
    """
    log.info(f"Streaming mode started — delay={delay}s per row")

    try:
        wb          = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
    except Exception as e:
        log.error(f"Failed to open workbook: {e}")
        return

    total_rows   = 0
    total_events = 0

    for sheet_name in sheet_names:
        log.info(f"[{sheet_name}] Starting stream...")

        # Clear any existing data for this session
        clear_session_data(sheet_name)

        # ── Reset trackers per session ────────────────────────────
        # Each sheet is an independent test case. Shared trackers
        # would cause state from one session to bleed into the next
        # (e.g. OUTSIDE at end of session 1 → Filter 3 discards the
        # Exit-before-entry in session 2).
        epc_trackers:  Dict[str, EPCTracker]      = {}
        mov_trackers:  Dict[str, MovementTracker]  = {}
        window_buffer: Dict[tuple, list]           = defaultdict(list)

        # Accumulate events in memory for end-of-session analytics
        session_events_list = []

        for raw_row in stream_rfid_excel(file_path, sheet_name, delay=delay):

            # Stage 1: persist raw read immediately
            try:
                ingest_single_row(sheet_name, raw_row)
                total_rows += 1
            except Exception as e:
                log.error(f"[{sheet_name}] ingest_single_row failed: {e}")
                continue

            # Stage 2: preprocess
            row = preprocess_row(raw_row)
            if row is None:
                continue

            epc = row["epc"]
            key = (epc, row["device"], row["direction"])

            if epc not in epc_trackers:
                epc_trackers[epc] = EPCTracker(epc)
                mov_trackers[epc] = MovementTracker(epc)

            # Stage 3: flush stale buffers — detect confirmed events
            confirmed = flush_stale_buffers(
                window_buffer, epc_trackers, mov_trackers, row["t0"]
            )

            # Stage 4: write each confirmed event immediately
            for event in confirmed:
                try:
                    write_event_to_db(sheet_name, event)
                    session_events_list.append(event)
                    total_events += 1
                    log.info(
                        f"[{sheet_name}] EVENT: {event.get('event')} "
                        f"t0={event.get('t0')} door={event.get('door','')}"
                    )
                except Exception as e:
                    log.error(f"[{sheet_name}] write_event_to_db failed: {e}")

            window_buffer[key].append(row)

        # ── End of session: flush remaining open buffers ──────────
        log.info(f"[{sheet_name}] Stream complete. Flushing remaining buffers...")

        remaining = sorted(
            [aggregate_window(buf) for buf in window_buffer.values() if buf],
            key=lambda x: x["t0_start"]
        )
        for agg in remaining:
            epc = agg["epc"]
            building_events = epc_trackers[epc].process_event(agg)
            for e in building_events:
                write_event_to_db(sheet_name, e)
                session_events_list.append(e)
                total_events += 1
                if e["event"] == "EXIT":
                    mov_trackers[epc].reset()
            mov_event = mov_trackers[epc].process_event(
                agg, epc_trackers[epc].state
            )
            if mov_event:
                write_event_to_db(sheet_name, mov_event)
                session_events_list.append(mov_event)
                total_events += 1

        for epc, tracker in epc_trackers.items():
            exit_event = tracker.flush_pending_exit()
            if exit_event:
                write_event_to_db(sheet_name, exit_event)
                session_events_list.append(exit_event)
                total_events += 1

        for epc, tracker in epc_trackers.items():
            if tracker.state in ("INSIDE", "UNKNOWN"):
                ended = {
                    "epc":   epc,
                    "event": "SESSION_ENDED_INSIDE",
                    "t0":    None,
                    "note":  "Stream ended before exit was detected",
                }
                write_event_to_db(sheet_name, ended)
                session_events_list.append(ended)

        # ── Post-session analytics ────────────────────────────────
        # Run anomaly detection and data quality on the completed session.
        # We use the in-memory session_events_list (which has "event" key)
        # rather than reading back from DB (which has "event_type" key)
        # so detect_anomalies() receives the format it expects.
        log.info(f"[{sheet_name}] Running post-session analytics...")
        try:
            db_rows   = get_raw_rows(sheet_name)
            anomalies = detect_anomalies(session_events_list, db_rows)
            quality   = compute_data_quality(session_events_list, db_rows)

            insert_anomalies(sheet_name, anomalies)
            update_session_quality(sheet_name, quality)

            if anomalies:
                from database import get_connection
                with get_connection() as conn:
                    conn.execute(
                        "UPDATE sessions SET has_anomaly = 1 WHERE session_id = ?",
                        (sheet_name,)
                    )

            log.info(
                f"[{sheet_name}] {len(anomalies)} anomalies. "
                f"Ghost ratio: {quality.get('ghost_read_ratio')}  "
                f"Entry signal: {quality.get('entry_rssi_strength')}"
            )

        except Exception as e:
            log.error(f"[{sheet_name}] Post-session analytics failed: {e}")

        log.info(f"[{sheet_name}] ✓ Session complete.")

    log.info(
        f"Streaming finished — "
        f"{total_rows} rows ingested, "
        f"{total_events} events detected."
    )


# ═══════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="RFID Pipeline")
    parser.add_argument(
        "--stream",
        action="store_true",
        help="Run in streaming mode (row by row with delay)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.02,
        help="Seconds between rows in streaming mode (default: 0.02)",
    )
    args = parser.parse_args()

    setup_logging()
    log.info("Pipeline started.")

    try:
        create_schema()
    except Exception as e:
        log.error(f"Failed to create DB schema: {e}")
        raise SystemExit(1)

    if args.stream:
        # ── Streaming mode ────────────────────────────────────────
        log.info(f"Mode: STREAMING  delay={args.delay}s per row")
        log.info("Open the dashboard in a separate terminal:")
        log.info("  streamlit run dashboard/app.py")
        print(f"\n{'═'*60}")
        print(f"  RFID Pipeline — Streaming Mode")
        print(f"  delay={args.delay}s per row")
        print(f"  Open dashboard: streamlit run dashboard/app.py")
        print(f"{'═'*60}\n")
        run_streaming(str(EXCEL_PATH), delay=args.delay)

    else:
        # ── Batch mode ────────────────────────────────────────────
        log.info("Mode: BATCH (full speed)")
        results = run_all_sheets(EXCEL_PATH)

        print(f"\n{'═'*60}")
        print(f"  SUMMARY")
        print(f"{'═'*60}")
        for session, events in results.items():
            entries = sum(1 for e in events if e["event"] == "ENTRY")
            exits   = sum(1 for e in events if e["event"] == "EXIT")
            flags   = sum(1 for e in events if e["event"] == "SESSION_ENDED_INSIDE")
            print(f"  {session:<35}  entries={entries}  exits={exits}  flags={flags}")

        print(f"\n  Database: {PROJECT_ROOT / 'data' / 'rfid_events.db'}")
        print(f"  Log file: {LOG_PATH}")
        log.info(f"Batch pipeline finished. {len(results)} sessions processed.")